import tkinter as tk
from PIL import ImageTk, Image
import os
import openpyxl
import datetime
import time

from tkinter import *


# from tkinter import ttk


class windows(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.wm_title("Test Application")
        self.configure(background='yellow')


class atlets():
    def __init__(self, window):
        self.atlets_now = tk.Label(text=f'Hello', fg="black", bg="pink", width=50, height=10,
                                   font=("Times New Roman", 14), justify="left", )
        self.atlets_next = tk.Label(text=f'Hello', fg="black", bg="red", width=50, height=10,
                                    font=("Times New Roman", 14), justify="left", )
        self.atlets_now.pack(padx=10, pady=10)
        self.atlets_next.pack(padx=10, pady=10)

    def update_atlets(self, window, sheet):
        for i in range(12, 50, 5):
            self.atlets_now["text"] = f'\nЗаход:{sheet.cell(row=i, column=2).value} \
                \nСейчас на поле следующие атлеты: \n\t{str(sheet.cell(row=i + 1, column=3).value)}\
                                                   \n\t{str(sheet.cell(row=i + 2, column=3).value)}\
                                                   \n\t{str(sheet.cell(row=i + 3, column=3).value)}\
                                                   \n\t{str(sheet.cell(row=i + 4, column=3).value)}'
            self.atlets_next["text"] = f'Old {i}'
            time.sleep(5)
            window.update()
        window.mainloop()

    def get_list_of_athletes(self):
        wb = openpyxl.load_workbook(filename='Тайминг 23 апреля.xlsx', data_only=True)
        sheet = wb["Лист1"]
        return sheet


class Logo():
    def __init__(self, window):
        canvas = Canvas(window, width=500, height=500)
        canvas.pack()
        img = PhotoImage(file='attachment.png')
        canvas.create_image(10, 10, anchor=NW, image=img)
        window.mainloop()



if __name__ == "__main__":
    testObj = windows()
    Logo = Logo(testObj)
    l = atlets(testObj)
    sheet = l.get_list_of_athletes()
    l.update_atlets(testObj, sheet)


    # sheet=get_workbook()
    # x=12
    # print(f'\nЗаход:{sheet.cell(row=x, column=2).value} \nСейчас на поле следующие атлеты: \n\t{str(sheet.cell(row=x, column=3).value)}')