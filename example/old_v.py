import tkinter as tk
from PIL import ImageTk, Image
import os
import openpyxl
import datetime
import time


def work_book():
    wb = openpyxl.load_workbook(filename='Тайминг 22 апреля.xlsx', data_only=True)
    sheet=wb["Лист1"]
    return sheet


def welcome():
    i = 12
    sheet = work_book()
    window = tk.Tk()
    window.title("Тайминг соревнований")

    label = tk.Label(
            text=f'\nЗаход:{sheet.cell(row=i, column=2).value} \nСейчас на поле следующие атлеты: \n\t{str(sheet.cell(row=i, column=3).value)}\n\t{str(sheet.cell(row=i + 1, column=3).value)}\n\t{str(sheet.cell(row=i + 2, column=3).value)}',
            fg="black",
            bg="white",
            width=40,
            height=5,
            font=("Times New Roman", 14),
            justify="left",

    )
        # кнопка закрытия приложения
    exit_button = tk.Button(window, text="Exit", command=window.destroy)

        # подгрузка изображения
    # img = Image.open("1-9610-512.png")
    # img = img.resize((200,200), Image.ANTIALIAS)
    # my_img = ImageTk.PhotoImage(img)
    # panel = tk.Label(window, image = my_img)

        # #распоковка виджетов в окне
    # panel.pack(fill=tk.BOTH, side=tk.LEFT, expand=False)
    label.pack(fill=tk.BOTH, side=tk.LEFT, expand=True, )
    exit_button.pack(pady=40)

        # изменение текстового значения
    for x in range(12, 60, 5):
        label[
                "text"] = f'\nЗаход:{sheet.cell(row=x, column=2).value} \nСейчас на поле следующие атлеты: \n\t{str(sheet.cell(row=x, column=3).value)}'
        time.sleep(5)
        window.update()
    window.mainloop()

if __name__ == "__main__":
    welcome()
